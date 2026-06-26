import re
from pathlib import Path

from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models.deletion import ProtectedError

from orders.models import Cart, CartItem, Order, OrderItem
from payments.models import Payment
from products.models import (
    Brand,
    Category,
    Product,
    ProductImage,
    ProductReview,
    WishlistItem,
)


COLUMN_ALIASES = {
    "name": ("نام محصول",),
    "category": ("دسته بندی", "دسته‌بندی"),
    "brand": ("برند",),
    "price": ("قیمت",),
    "discount_price": (
        "قیمت تخفیفی در صورت وجود",
        "قیمت تخفیفی در صورت وجود وجودی",
    ),
    "stock": ("موجودی",),
    "age_group": ("گروه سنی",),
    "player_count": ("تعداد بازیکن",),
    "play_time": ("مدت زمان بازی",),
    "short_description": ("توضیح کوتاه",),
    "description": ("توضیح کامل",),
    "feature": ("ویژگی", "ویژگی**"),
    "main_image": ("نام تصویر اصلی",),
    "gallery_images": ("نام تصاویر بیشتر",),
}

IMAGE_FOLDER_ALIASES = {
    "هوش چین یک تکه": "هوش چین",
    "لگو بتمن": "لگو بتمن",
    "برس بیرمنگهام": "برس بیرمین",
    "مونوپولی فکر اوران": "مونوپولی کلاسیک",
    "گربه انفجاری": "گربه انفجاری",
    "چکش شکاری": "چکش شکاری",
    "اوسا بنا 1+": "اوسا بنا",
    "خانه جنگلی": "خانه جنگلی",
    "روبیک 3 در 3": "روبیک سه در سه کای وای",
}

PERSIAN_WORD_NUMBERS = {
    "صفر": 0,
    "یک": 1,
    "دو": 2,
    "سه": 3,
    "چهار": 4,
    "پنج": 5,
    "شش": 6,
    "هفت": 7,
    "هشت": 8,
    "نه": 9,
    "ده": 10,
    "یازده": 11,
    "دوازده": 12,
}

PERSIAN_TRANSLITERATION = {
    "آ": "a",
    "ا": "a",
    "أ": "a",
    "إ": "a",
    "ب": "b",
    "پ": "p",
    "ت": "t",
    "ث": "s",
    "ج": "j",
    "چ": "ch",
    "ح": "h",
    "خ": "kh",
    "د": "d",
    "ذ": "z",
    "ر": "r",
    "ز": "z",
    "ژ": "zh",
    "س": "s",
    "ش": "sh",
    "ص": "s",
    "ض": "z",
    "ط": "t",
    "ظ": "z",
    "ع": "",
    "غ": "gh",
    "ف": "f",
    "ق": "gh",
    "ک": "k",
    "ك": "k",
    "گ": "g",
    "ل": "l",
    "م": "m",
    "ن": "n",
    "و": "v",
    "ؤ": "v",
    "ه": "h",
    "ة": "h",
    "ی": "y",
    "ي": "y",
    "ئ": "y",
}

PERSIAN_DIGITS = str.maketrans(
    "۰۱۲۳۴۵۶۷۸۹" "٠١٢٣٤٥٦٧٨٩",
    "0123456789" "0123456789",
)
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

PRODUCT_NAME_MAX_LENGTH = 255
PRODUCT_SLUG_MAX_LENGTH = 275
PRODUCT_SHORT_DESCRIPTION_MAX_LENGTH = 500
PRODUCT_SKU_MAX_LENGTH = 100
PRODUCT_AGE_GROUP_MAX_LENGTH = 20
PRODUCT_GENDER_MAX_LENGTH = 10


class Command(BaseCommand):
    help = "Import real IpakToys products from an Excel file and local images."

    def add_arguments(self, parser):
        default_base_dir = Path("import_data")
        parser.add_argument(
            "--excel",
            default=str(default_base_dir / "products.xlsx"),
            help="Path to the products Excel file.",
        )
        parser.add_argument(
            "--images-dir",
            default=str(default_base_dir / "product_images"),
            help="Path to extracted product image folders.",
        )
        parser.add_argument(
            "--clear-existing",
            action="store_true",
            help="Delete existing products before importing. Use only after backup.",
        )
        parser.add_argument(
            "--clear-dependent-demo-data",
            action="store_true",
            help=(
                "With --clear-existing, delete development dependent data "
                "before deleting products. This removes payments, orders, "
                "carts, wishlist items, reviews and product images."
            ),
        )

    def handle(self, *args, **options):
        if options["clear_dependent_demo_data"] and not options["clear_existing"]:
            raise CommandError(
                "--clear-dependent-demo-data requires --clear-existing. "
                "This guard prevents accidental deletion of development data."
            )

        excel_path = Path(options["excel"]).resolve()
        images_dir = Path(options["images_dir"]).resolve()

        if not excel_path.exists():
            raise CommandError(f"Excel file not found: {excel_path}")
        if not images_dir.exists():
            raise CommandError(f"Images directory not found: {images_dir}")

        rows = self.load_rows(excel_path)
        image_folders = self.build_image_folder_index(images_dir)
        summary = ImportSummary()

        if options["clear_existing"]:
            self.clear_existing_products(
                clear_dependent_demo_data=options["clear_dependent_demo_data"]
            )

        for row_number, row in rows:
            try:
                with transaction.atomic():
                    result = self.import_row(row_number, row, image_folders)
            except Exception as exc:
                product_name = get_column(row, "name") or f"row {row_number}"
                result = ImportResult(
                    skipped=True,
                    row_number=row_number,
                    product_name=product_name,
                    warnings=(f"Import failed: {exc}",),
                )
            summary.add(result)

        summary.print(self)

    def load_rows(self, excel_path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required. Install backend requirements first."
            ) from exc

        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        headers = [clean_text(cell.value) for cell in sheet[1]]
        rows = []

        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row = {
                header: value
                for header, value in zip(headers, values, strict=False)
                if header
            }
            rows.append((row_number, row))

        return rows

    def import_row(self, row_number, row, image_folders):
        warnings = []
        raw_name = get_column(row, "name")
        name = truncate_field(
            raw_name,
            PRODUCT_NAME_MAX_LENGTH,
            row_number,
            raw_name or f"row {row_number}",
            "name",
            warnings,
        )
        category_name = get_column(row, "category")
        price = parse_integer(get_column(row, "price"))
        stock = parse_inventory(get_column(row, "stock"))

        if not name or not category_name or price is None or stock is None:
            return ImportResult(
                skipped=True,
                row_number=row_number,
                product_name=name or f"row {row_number}",
                warnings=("Missing required product name/category/price/stock.",),
            )

        brand_name = get_column(row, "brand") or "IpakToys"
        discount_price = parse_integer(get_column(row, "discount_price"))
        if discount_price is not None and discount_price >= price:
            warnings.append("Discount price was ignored because it is not below price.")
            discount_price = None

        category = upsert_category(category_name)
        brand = upsert_brand(brand_name)
        existing_product = Product.objects.filter(name=name).first()
        base_slug = truncate_field(
            make_slug(name),
            PRODUCT_SLUG_MAX_LENGTH,
            row_number,
            name,
            "slug",
            warnings,
        )
        product = existing_product or Product.objects.filter(slug=base_slug).first()
        slug = product.slug if product else make_unique_product_slug(base_slug)
        sku = product.sku if product else make_unique_sku(slug)
        sku = truncate_field(
            sku,
            PRODUCT_SKU_MAX_LENGTH,
            row_number,
            name,
            "sku",
            warnings,
        )
        age_group = truncate_field(
            parse_age_group(get_column(row, "age_group")),
            PRODUCT_AGE_GROUP_MAX_LENGTH,
            row_number,
            name,
            "age_group",
            warnings,
        )
        gender = truncate_field(
            Product.Gender.UNISEX,
            PRODUCT_GENDER_MAX_LENGTH,
            row_number,
            name,
            "gender",
            warnings,
        )
        # Product.short_description is a CharField(max_length=500). Detailed
        # Excel content such as features, player count, play time and age group
        # belongs in Product.description, which is a TextField.
        short_description = truncate_field(
            get_column(row, "short_description") or name,
            PRODUCT_SHORT_DESCRIPTION_MAX_LENGTH,
            row_number,
            name,
            "short_description",
            warnings,
        )

        defaults = {
            "name": name,
            "slug": slug,
            "category": category,
            "brand": brand,
            "description": build_description(row),
            "short_description": short_description,
            "sku": sku,
            "price": price,
            "discount_price": discount_price,
            "stock": stock,
            "age_group": age_group,
            "gender": gender,
            "is_active": True,
            "is_featured": False,
        }

        if product:
            for field, value in defaults.items():
                setattr(product, field, value)
            product.save()
            created = False
        else:
            product = Product.objects.create(**defaults)
            created = True

        image_result = attach_product_images(product, image_folders)
        warnings.extend(image_result.warnings)

        return ImportResult(
            imported=created,
            updated=not created,
            row_number=row_number,
            product_name=name,
            missing_image=image_result.missing,
            warnings=tuple(warnings),
        )

    def build_image_folder_index(self, images_dir):
        folders = {}
        for path in images_dir.iterdir():
            if path.is_dir():
                folders[normalize_match_text(path.name)] = path
        return folders

    def clear_existing_products(self, clear_dependent_demo_data=False):
        if clear_dependent_demo_data:
            self.stdout.write(
                self.style.WARNING(
                    "DESTRUCTIVE ACTION: --clear-dependent-demo-data is enabled. "
                    "This will delete development payments, orders, carts, "
                    "wishlist items, reviews, product images and products."
                )
            )
            with transaction.atomic():
                deleted_counts = self.delete_dependent_demo_data()
                deleted_products, _ = Product.objects.all().delete()

            self.stdout.write(
                self.style.WARNING(
                    "Deleted dependent demo data: "
                    + ", ".join(
                        f"{label}={count}"
                        for label, count in deleted_counts.items()
                    )
                )
            )
            self.stdout.write(
                self.style.WARNING(
                    f"--clear-existing removed {deleted_products} product "
                    "records and related rows."
                )
            )
            return

        try:
            with transaction.atomic():
                deleted_count, _ = Product.objects.all().delete()
        except ProtectedError as exc:
            protected_count = len(exc.protected_objects)
            raise CommandError(
                "--clear-existing could not delete products because some "
                "products are referenced by existing order items. For "
                "production safety, this command will not delete orders, "
                "payments or customer data unless you explicitly pass "
                "--clear-dependent-demo-data. Take a database backup first "
                "if this is not a disposable local database, then run: "
                "python manage.py import_real_products --clear-existing "
                "--clear-dependent-demo-data. "
                f"Protected related objects: {protected_count}."
            ) from exc

        self.stdout.write(
            self.style.WARNING(
                f"--clear-existing removed {deleted_count} product records "
                "and related rows."
            )
        )

    def delete_dependent_demo_data(self):
        deleted_counts = {}

        # Payment.order and OrderItem.product use PROTECT, so payments and order
        # items must go before orders/products in this explicit demo reset path.
        deleted_counts["payments"] = Payment.objects.all().delete()[0]
        deleted_counts["order_items"] = OrderItem.objects.all().delete()[0]
        deleted_counts["orders"] = Order.objects.all().delete()[0]
        deleted_counts["cart_items"] = CartItem.objects.all().delete()[0]
        deleted_counts["carts"] = Cart.objects.all().delete()[0]
        deleted_counts["wishlist_items"] = WishlistItem.objects.all().delete()[0]
        deleted_counts["reviews"] = ProductReview.objects.all().delete()[0]
        deleted_counts["product_images"] = ProductImage.objects.all().delete()[0]

        return deleted_counts


class ImportSummary:
    def __init__(self):
        self.imported = 0
        self.updated = 0
        self.skipped = []
        self.missing_images = []
        self.warnings = []

    def add(self, result):
        if result.imported:
            self.imported += 1
        if result.updated:
            self.updated += 1
        if result.skipped:
            self.skipped.append(result)
        if result.missing_image:
            self.missing_images.append(result.product_name)
        self.warnings.extend(
            f"Row {result.row_number} ({result.product_name}): {warning}"
            for warning in result.warnings
        )

    def print(self, command):
        command.stdout.write(command.style.SUCCESS("Import completed."))
        command.stdout.write(f"Imported products: {self.imported}")
        command.stdout.write(f"Updated products: {self.updated}")
        command.stdout.write(f"Skipped rows: {len(self.skipped)}")
        command.stdout.write(f"Missing images: {len(self.missing_images)}")
        command.stdout.write(f"Validation warnings: {len(self.warnings)}")

        if self.skipped:
            command.stdout.write(command.style.WARNING("Skipped rows:"))
            for result in self.skipped:
                command.stdout.write(
                    f"  - Row {result.row_number}: {result.product_name}"
                )

        if self.missing_images:
            command.stdout.write(command.style.WARNING("Missing image folders:"))
            for name in self.missing_images:
                command.stdout.write(f"  - {name}")

        if self.warnings:
            command.stdout.write(command.style.WARNING("Warnings:"))
            for warning in self.warnings:
                command.stdout.write(f"  - {warning}")


class ImportResult:
    def __init__(
        self,
        product_name,
        row_number,
        imported=False,
        updated=False,
        skipped=False,
        missing_image=False,
        warnings=(),
    ):
        self.product_name = product_name
        self.row_number = row_number
        self.imported = imported
        self.updated = updated
        self.skipped = skipped
        self.missing_image = missing_image
        self.warnings = warnings


class ImageAttachResult:
    def __init__(self, missing=False, warnings=()):
        self.missing = missing
        self.warnings = warnings


def get_column(row, key):
    for column_name in COLUMN_ALIASES[key]:
        value = row.get(column_name)
        if value is not None and clean_text(value):
            return clean_text(value)
    return ""


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def truncate_text(value, max_length, suffix="..."):
    text = clean_text(value)
    if len(text) <= max_length:
        return text

    if max_length <= 0:
        return ""

    if suffix and max_length > len(suffix):
        return text[: max_length - len(suffix)].rstrip() + suffix

    return text[:max_length].rstrip()


def truncate_field(value, max_length, row_number, product_name, field_name, warnings):
    text = clean_text(value)
    suffix = "" if field_name in {"slug", "sku", "age_group", "gender"} else "..."
    truncated = truncate_text(text, max_length, suffix=suffix)

    if len(text) > max_length:
        warnings.append(
            f"truncated {field_name} from {len(text)} to "
            f"{len(truncated)} characters."
        )

    return truncated


def parse_integer(value):
    text = clean_text(value)
    if not text:
        return None

    normalized = text.translate(PERSIAN_DIGITS).replace(",", "")
    if "تخفیف ندار" in normalized:
        return None

    match = re.search(r"\d+", normalized)
    if not match:
        return None
    return int(match.group())


def parse_inventory(value):
    text = clean_text(value)
    if not text:
        return None

    normalized = text.translate(PERSIAN_DIGITS).replace(",", "")
    if normalized in PERSIAN_WORD_NUMBERS:
        return PERSIAN_WORD_NUMBERS[normalized]

    match = re.search(r"\d+", normalized)
    if match:
        return int(match.group())

    return None


def parse_age_group(value):
    text = clean_text(value).translate(PERSIAN_DIGITS)
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    min_age = min(numbers) if numbers else 12

    if min_age <= 2:
        return Product.AgeGroup.ZERO_TO_TWO
    if min_age <= 5:
        return Product.AgeGroup.THREE_TO_FIVE
    if min_age <= 8:
        return Product.AgeGroup.SIX_TO_EIGHT
    if min_age <= 12:
        return Product.AgeGroup.NINE_TO_TWELVE
    return Product.AgeGroup.TWELVE_PLUS


def build_description(row):
    sections = []
    description = get_column(row, "description")
    short_description = get_column(row, "short_description")
    feature = get_column(row, "feature")
    player_count = get_column(row, "player_count")
    play_time = get_column(row, "play_time")
    age_group = get_column(row, "age_group")

    if description:
        sections.append(description)
    elif short_description:
        sections.append(short_description)

    specs = []
    if feature:
        specs.append(f"ویژگی: {feature}")
    if age_group:
        specs.append(f"گروه سنی: {age_group}")
    if player_count:
        specs.append(f"تعداد بازیکن: {player_count}")
    if play_time:
        specs.append(f"مدت زمان بازی: {play_time}")

    if specs:
        sections.append("\n".join(specs))

    return "\n\n".join(sections) or "محصول فروشگاه IpakToys"


def upsert_category(name):
    slug = make_unique_category_slug(make_slug(name), name)
    category, _ = Category.objects.update_or_create(
        name=name,
        defaults={
            "slug": slug,
            "description": f"محصولات دسته {name} در IpakToys",
            "is_active": True,
        },
    )
    return category


def upsert_brand(name):
    slug = make_unique_brand_slug(make_slug(name), name)
    brand, _ = Brand.objects.update_or_create(
        name=name,
        defaults={
            "slug": slug,
            "description": f"محصولات برند {name} در IpakToys",
            "is_active": True,
        },
    )
    return brand


def make_unique_category_slug(base_slug, name):
    category = Category.objects.filter(name=name).first()
    return category.slug if category else make_unique_slug(Category, base_slug)


def make_unique_brand_slug(base_slug, name):
    brand = Brand.objects.filter(name=name).first()
    return brand.slug if brand else make_unique_slug(Brand, base_slug)


def make_unique_product_slug(base_slug):
    return make_unique_slug(Product, base_slug, PRODUCT_SLUG_MAX_LENGTH)


def make_unique_slug(model, base_slug, max_length=170):
    slug = truncate_text(base_slug, max_length, suffix="")
    counter = 2
    while model.objects.filter(slug=slug).exists():
        suffix = f"-{counter}"
        trimmed_base = truncate_text(base_slug, max_length - len(suffix), suffix="")
        slug = f"{trimmed_base}{suffix}"
        counter += 1
    return slug


def make_unique_sku(slug):
    base_sku = f"IPAK-{re.sub(r'[^A-Z0-9]+', '-', slug.upper()).strip('-')}"
    base_sku = base_sku[:90].rstrip("-") or "IPAK-PRODUCT"
    sku = base_sku
    counter = 2
    while Product.objects.filter(sku=sku).exists():
        sku = f"{base_sku[:85]}-{counter}"
        counter += 1
    return sku


def make_slug(value):
    text = clean_text(value).translate(PERSIAN_DIGITS).lower()
    parts = []
    for char in text:
        if char.isascii() and char.isalnum():
            parts.append(char)
        elif char in PERSIAN_TRANSLITERATION:
            parts.append(PERSIAN_TRANSLITERATION[char])
        else:
            parts.append("-")

    slug = re.sub(r"-+", "-", "".join(parts)).strip("-")
    return slug or "ipak-product"


def attach_product_images(product, image_folders):
    folder = find_image_folder(product.name, image_folders)
    if folder is None:
        return ImageAttachResult(missing=True)

    image_paths = sorted(
        path
        for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    )
    if not image_paths:
        return ImageAttachResult(
            missing=True,
            warnings=(
                f"Image folder exists but contains no supported images: {folder}",
            ),
        )

    product.images.all().delete()
    warnings = []
    for index, image_path in enumerate(image_paths):
        try:
            with image_path.open("rb") as image_file:
                image = ProductImage(
                    product=product,
                    alt_text=product.name,
                    is_main=index == 0,
                )
                image.image.save(
                    f"imported/{product.slug}/{image_path.name}",
                    File(image_file),
                    save=True,
                )
        except OSError as exc:
            warnings.append(f"Could not attach image {image_path.name}: {exc}")

    return ImageAttachResult(warnings=tuple(warnings))


def find_image_folder(product_name, image_folders):
    alias = IMAGE_FOLDER_ALIASES.get(product_name, product_name)
    normalized_alias = normalize_match_text(alias)
    if normalized_alias in image_folders:
        return image_folders[normalized_alias]

    normalized_name = normalize_match_text(product_name)
    for folder_key, folder_path in image_folders.items():
        if folder_key in normalized_name or normalized_name in folder_key:
            return folder_path

    return None


def normalize_match_text(value):
    return (
        clean_text(value)
        .translate(PERSIAN_DIGITS)
        .replace("ي", "ی")
        .replace("ك", "ک")
        .replace("ۀ", "ه")
        .replace("ة", "ه")
        .replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace(" ", "")
        .replace("-", "")
        .lower()
    )
